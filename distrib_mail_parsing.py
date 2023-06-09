import asyncio
import base64
import email
import imaplib
import quopri

from openpyxl.reader.excel import load_workbook
from email.header import decode_header

from openpyxl.workbook import Workbook

from config import hidden_vars
from core_func import date_out, android_profit

from core_vars import y, sqlite_connection


def mail_connect():
    global con
    con = imaplib.IMAP4_SSL("imap.mail.ru")
    con.login(hidden_vars.mail_connect.mailbox, hidden_vars.mail_connect.mail_pass)
    con.select(hidden_vars.mail_connect.mail_path)
    result, data_connect = con.search(None, "UNSEEN")
    msg_list = list(data_connect[0].decode('UTF-8').split())
    if msg_list:
        return msg_list


def mail_processing(msg_list):
    to_be_write_into_db = list()
    for i in msg_list:
        result, data = con.fetch(i, "(RFC822)")
        msg = email.message_from_bytes(data[0][1])
        date_time_letter = date_out(email.utils.parsedate_to_datetime(msg["Date"]))
        subject = decode_header(msg["Subject"])[0][0].decode()
        if hidden_vars.mail_connect.subject_keywords_xls in subject:
            if msg.is_multipart():
                for part in msg.walk():
                    if part.get_content_maintype() == 'multipart' or part.get('Content-Disposition') is None:
                        continue
                    filename = part.get_filename()
                    transfer_encoding = part.get_all('Content-Transfer-Encoding')
                    if transfer_encoding and transfer_encoding[0] == 'base64':
                        filename_parts = filename.split('?')
                        filename = base64.b64decode(filename_parts[3]).decode(filename_parts[1])
                        if '.xlsx' or '.xls' in filename:
                            with open(
                                    f'shippers/{hidden_vars.mail_connect.mail_path}/{filename}', 'wb'
                            ) as new_file:
                                new_file.write(part.get_payload(decode=True))
                            y.upload(f'shippers/{hidden_vars.mail_connect.mail_path}/{filename}',
                                     f'/shippers/Mobex/{filename}', overwrite=True)
                            to_be_write_into_db.append([filename, date_time_letter])
        if hidden_vars.mail_connect.subject_keywords_apple in subject:
            for part in msg.walk():
                if part.get_content_maintype() == 'text' and part.get_content_subtype() == 'plain':
                    transfer_encoding = part.get_all('Content-Transfer-Encoding')
                    if transfer_encoding and transfer_encoding[0] == 'base64':
                        text = base64.b64decode(part.get_payload()).decode()
                        to_be_write_into_db = add_excel_data(text, date_time_letter)
                    if transfer_encoding and transfer_encoding[0] == 'quoted-printable':
                        text = quopri.decodestring(part.get_payload()).decode('utf-8')
                        to_be_write_into_db = add_excel_data(text, date_time_letter)
        return to_be_write_into_db


#
def add_excel_data(text_data, date):
    to_be_write_into_db = list()
    in_list = list()
    for line in text_data.split('\n'):
        if len(line) > 5 and '-' in line:
            in_list.append(line.partition('-'))
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"
    ws.append(['Наименование', 'Цена', 'Заказ'])
    for row in in_list:
        ws.append([row[0].strip(), int(row[2].replace(' ', ''))])
    filename = hidden_vars.mail_connect.subject_keywords_apple.replace(' ', '_') + '_' + date[:10] + '.xlsx'
    wb.save(f'shippers/{hidden_vars.mail_connect.mail_path}/{filename}')
    y.upload(f'shippers/{hidden_vars.mail_connect.mail_path}/{filename}', f'/shippers/Mobex/{filename}', overwrite=True)
    to_be_write_into_db.append([filename, date])
    return to_be_write_into_db


def check_data_in_distributor(date, distributor_price_list):
    sqlite_cur = sqlite_connection.cursor()
    sqlite_cur.execute(
        f"SELECT DATE FROM {distributor_price_list} WHERE DATE = '{str(date)}' "
        f"GROUP BY DATE"
    )
    result = sqlite_cur.fetchall()
    try:
        return result
    except TypeError:
        return None


def from_xls_into_db(data_list):
    price_list_name = str()
    for price in range(len(data_list)):
        if 'к.xlsx' in data_list[price][0]:
            price_list_name = 'optmobex_xiaomi'
        if 'sams.xlsx' in data_list[price][0]:
            price_list_name = 'optmobex_samsung'
        if 'Apple' in data_list[price][0]:
            price_list_name = 'optmobex_apple'
        result_checking = check_data_in_distributor(data_list[price][1], price_list_name)
        if result_checking:
            print(f'В БД {price_list_name} уже есть такой прайс:', *data_list[price])
            pass
        else:
            print(f'Заношу в БД {price_list_name} такой прайс:', *data_list[price])
            price_list = list()
            wb = load_workbook('shippers/' + hidden_vars.mail_connect.mail_path + '/' + data_list[price][0])
            ws = wb["Лист1"]
            rows = ws.max_row
            cols = ws.max_column - 1
            sqlite_cur = sqlite_connection.cursor()
            for i in range(2, rows):
                string = str()
                for j in range(1, cols + 1):
                    cell = ws.cell(row=i, column=j)
                    string = string + str(cell.value) + ' '
                price_list.append(string.strip(' ').split(' '))
            for k in price_list:
                product_name = ' '.join(k[:-1])
                input_price = int(k[-1])
                out_price = android_profit(int(k[-1]))
                sqlite_cur.execute(f"INSERT INTO {price_list_name} VALUES "
                                   f"('{data_list[price][1]}', "
                                   f"'{product_name}', "
                                   f"'{input_price}', "
                                   f"'{out_price}')")
            sqlite_connection.commit()
            print('Запись:', data_list[price][0], 'завершена')
            y.upload('db/cifrotech_db', '/shippers/cifrotech_db', overwrite=True)


async def mail_parsing():
    while True:
        response = mail_connect()
        if response:
            prepare_letters = mail_processing(response)
            if prepare_letters:
                from_xls_into_db(prepare_letters)
        await asyncio.sleep(10)
