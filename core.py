import imaplib
import os
from datetime import datetime

import requests
import yadisk
from openpyxl.reader.excel import load_workbook
import pytz

from config import hidden_vars

samsung_xlsx_list = list()
xiaomi_xlsx_list = list()
apple_txt_list = list()


def date_out(date):
    m_date = datetime.strptime(str(date), "%Y-%m-%d %H:%M:%S%z")
    tz = pytz.timezone("Etc/GMT-3")
    m_date_utc3 = tz.normalize(m_date.astimezone(tz))
    out_date = m_date_utc3.strftime("%d-%m-%Y %H:%M")
    return out_date


def ya_time_converter(time):
    a = time
    tz = pytz.timezone("Etc/GMT-3")
    date = tz.normalize(a.astimezone(tz))
    return date.date().strftime("%d %m %Y")


days = ({0: "Понедельник"},
        {1: "Вторник"},
        {2: "Среда"},
        {3: "Четверг"},
        {4: "Пятница"},
        {5: "Суббота"},
        {6: "Воскресенье"})

month = ({1: "Января"},
         {2: "Февраля"},
         {3: "Марта"},
         {4: "Апреля"},
         {5: "Мая"},
         {6: "Июня"},
         {7: "Июля"},
         {8: "Августа"},
         {9: "Сентября"},
         {10: "Октября"},
         {11: "Ноября"},
         {12: "Декабря"})

xiaomi_del_list = [' EU ', ' RU ', ' RUСТБ ', 'Xiaomi ', ' JP ']
samsung_del_list = [' KZ ', ' AE ', ' AH ', ' EU ', ' RUTH ',
                    ' Simfree ', ' KR ', ' KZEU ', ' CNINAH ',
                    'INAHMY ', ' IN ', ' KZAEZA ', ' TH ', ' RU ', ' KZAEEU ', '  ', 'Samsung Galaxy ']


def text_file_order_list(price_file):
    with open(price_file, 'r', encoding='utf8') as f:
        file = f.read()
    our_price = int()
    apple_ord = dict()
    apple_order_list = file.split('\n')
    for i in apple_order_list:
        k = i.split(' ')
        if int(k[-1]) in range(10000, 20000):
            our_price = int(k[-1]) + 1500
        if int(k[-1]) in range(20000, 30000):
            our_price = int(k[-1]) + 2500
        if int(k[-1]) in range(30000, 40000):
            our_price = int(k[-1]) + 3000
        if int(k[-1]) in range(40000, 50000):
            our_price = int(k[-1]) + 3500
        if int(k[-1]) in range(50000, 100000):
            our_price = int(k[-1]) + 5000
        if int(k[-1]) in range(100000, 300000):
            our_price = int(k[-1]) + 8000
        y = [(" ".join(k[:-1]), our_price)]
        apple_ord.update(y)
        y.clear()
    return apple_ord


def excel_order_list(file_list, del_list):
    order_price = int()
    price_list = list()
    xiaomi_ord = dict()
    wb = load_workbook('shippers/' + str(file_list))
    ws = wb["Лист1"]
    rows = ws.max_row
    cols = ws.max_column - 1
    for i in range(2, rows):
        string = str()
        for j in range(1, cols + 1):
            cell = ws.cell(row=i, column=j)
            string = string.replace(",", "").replace("/", "")
            for t in del_list:
                string = string.replace(t, ' ')
            string = string + str(cell.value) + ' '
        price_list.append(string.strip(' ').split(' '))
    price_list.sort(key=lambda arr: int(arr[-1]))
    for k in price_list:
        if int(k[-1]) in range(4000, 8000):
            order_price = int(k[-1]) + 1500
        if int(k[-1]) in range(8000, 15000):
            order_price = int(k[-1]) + 2500
        if int(k[-1]) in range(15000, 20000):
            order_price = int(k[-1]) + 3000
        if int(k[-1]) in range(20000, 30000):
            order_price = int(k[-1]) + 3500
        if int(k[-1]) in range(30000, 50000):
            order_price = int(k[-1]) + 5000
        if int(k[-1]) in range(50000, 300000):
            order_price = int(k[-1]) + 8000
        y = [(" ".join(k[:-1]), order_price)]
        xiaomi_ord.update(y)
    price_list.clear()
    for key, value in xiaomi_ord.items():
        price_list.append((str(key) + ' - ' + str(value)))
    return price_list


def actual_date():
    result = list()
    actual = open('shippers/actual.txt', encoding='ANSI')
    for i in actual:
        result.append(i.replace('\n', ' ').rstrip()[-10:])
    return result


# def update_shippers_data():
#     files_dict = dict()
#     y = yadisk.YaDisk(token=hidden_vars.misc_path.yadisk)
#     for i in list(y.listdir('shippers/apple')):
#         apple_txt_list.append(i['name'])
#         apple_txt_list.append(ya_time_converter(i['modified']))
#         apple_txt_list.append(i['file'])
#     if os.path.isfile('shippers/apple/' + str(apple_txt_list[0])):
#         pass
#     else:
#         download_txt = requests.get(apple_txt_list[2])
#         with open('shippers/apple/' + str(apple_txt_list[0]), 'wb') as f:
#             f.write(download_txt.content)
#         f.close()
#     for i in list(y.listdir('shippers')):
#         files_dict[i['name']] = [ya_time_converter(i['modified']), i['file']]
#     for i in list(files_dict.keys()):
#         if 'sams' in i:
#             samsung_xlsx_list.append(i)
#             samsung_xlsx_list.append(files_dict.get(i)[0])
#             samsung_xlsx_list.append(files_dict.get(i)[1])
#             if os.path.isfile('shippers/' + str(samsung_xlsx_list[0])):
#                 pass
#             else:
#                 download_response = requests.get(samsung_xlsx_list[2])
#                 with open('shippers/' + str(samsung_xlsx_list[0]), 'wb') as f:
#                     f.write(download_response.content)
#                 f.close()
#         if 'к.xlsx' in i:
#             xiaomi_xlsx_list.append(i)
#             xiaomi_xlsx_list.append(files_dict.get(i)[0])
#             xiaomi_xlsx_list.append(files_dict.get(i)[1])
#             if os.path.isfile('shippers/' + str(xiaomi_xlsx_list[0])):
#                 pass
#             else:
#                 download_response = requests.get(xiaomi_xlsx_list[2])
#                 with open('shippers/' + str(xiaomi_xlsx_list[0]), 'wb') as f:
#                     f.write(download_response.content)
#                 f.close()





def update_price_list_data():
    pass
